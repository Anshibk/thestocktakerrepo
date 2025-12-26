from __future__ import annotations

from collections import OrderedDict
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.category import CategoryGroup, SubCategory
from app.models.item import Item
from app.core.constants import CORE_INVENTORY_GROUPS_SET


EXPECTED_HEADERS = {
    "item name": "item_name",
    "group": "group_name",
    "sub category": "sub_category",
    "unit": "unit",
    "price": "price",
}


def _normalise_string(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _coerce_price(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as exc:  # noqa: BLE001
        raise ValueError("Price must be a valid number") from exc


def import_items(db: Session, payload: bytes, *, original_filename: str = "import.xlsx") -> dict[str, int]:
    try:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True)
    except (InvalidFileException, OSError) as exc:  # noqa: BLE001
        raise ValueError("Upload a valid .xlsx workbook") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The workbook is empty")

    header_row = rows[0]
    column_map: dict[int, str] = {}
    seen_headers: set[str] = set()
    for idx, raw_header in enumerate(header_row):
        header_name = _normalise_string(raw_header).lower()
        if not header_name:
            continue
        if header_name in EXPECTED_HEADERS and header_name not in seen_headers:
            column_map[idx] = EXPECTED_HEADERS[header_name]
            seen_headers.add(header_name)

    if set(EXPECTED_HEADERS.values()) - set(column_map.values()):
        raise ValueError(
            "The first row must contain headers for Item Name, Group, Sub Category, Unit, and Price"
        )

    deduped: "OrderedDict[str, dict[str, Any]]" = OrderedDict()
    skipped = 0

    for row in rows[1:]:
        if row is None:
            continue
        record: dict[str, Any] = {"source_row": row}
        for idx, key in column_map.items():
            record[key] = row[idx] if idx < len(row) else None

        item_name = _normalise_string(record.get("item_name"))
        group_name = _normalise_string(record.get("group_name"))
        unit = _normalise_string(record.get("unit"))
        sub_category = _normalise_string(record.get("sub_category"))

        if not item_name or not group_name or not unit:
            skipped += 1
            continue

        if group_name not in CORE_INVENTORY_GROUPS_SET:
            skipped += 1
            continue

        try:
            price = _coerce_price(record.get("price"))
        except ValueError:
            skipped += 1
            continue

        deduped[item_name.lower()] = {
            "item_name": item_name,
            "group_name": group_name,
            "sub_category": sub_category or None,
            "unit": unit,
            "price": price,
        }

    if not deduped:
        raise ValueError("No valid rows found in the spreadsheet")

    group_names = {row["group_name"].lower() for row in deduped.values()}
    groups = (
        db.query(CategoryGroup)
        .filter(func.lower(CategoryGroup.name).in_(group_names))
        .all()
    )
    group_lookup = {group.name.lower(): group for group in groups}
    missing_groups = sorted(group_names - set(group_lookup))
    if missing_groups:
        raise ValueError(
            "Inventory group is not recognised. Please use one of the official group names."
        )

    requested_subcategories: dict[tuple[UUID, str], str] = {}
    for row in deduped.values():
        sub_name = row.get("sub_category")
        if not sub_name:
            continue
        group = group_lookup[row["group_name"].lower()]
        key = (group.id, sub_name.lower())
        if key not in requested_subcategories:
            requested_subcategories[key] = sub_name

    existing_subcategories: dict[tuple[UUID, str], SubCategory] = {}
    if requested_subcategories:
        group_ids = {group_id for group_id, _ in requested_subcategories.keys()}
        existing = (
            db.query(SubCategory)
            .filter(SubCategory.group_id.in_(group_ids))
            .all()
        )
        for sub in existing:
            existing_subcategories[(sub.group_id, sub.name.lower())] = sub

        new_subs: list[SubCategory] = []
        for key, original_name in requested_subcategories.items():
            if key in existing_subcategories:
                continue
            group_id, _ = key
            sub = SubCategory(name=original_name, group_id=group_id)
            new_subs.append(sub)
            existing_subcategories[key] = sub
        if new_subs:
            db.add_all(new_subs)
            db.flush()

    item_names = list(deduped.keys())
    item_lookup: dict[str, Item] = {}
    if item_names:
        existing_items = (
            db.query(Item)
            .filter(func.lower(Item.name).in_(item_names))
            .all()
        )
        item_lookup = {item.name.lower(): item for item in existing_items}

    created = 0
    updated = 0

    for payload_row in deduped.values():
        group = group_lookup[payload_row["group_name"].lower()]
        sub_name = payload_row.get("sub_category")
        sub: SubCategory | None = None
        if sub_name:
            sub = existing_subcategories.get((group.id, sub_name.lower()))

        existing_item = item_lookup.get(payload_row["item_name"].lower())

        if existing_item:
            existing_item.unit = payload_row["unit"]
            existing_item.price = payload_row["price"]
            existing_item.category_id = sub.id if sub else None
            updated += 1
            continue

        item = Item(
            name=payload_row["item_name"],
            unit=payload_row["unit"],
            price=payload_row["price"],
            category_id=sub.id if sub else None,
        )
        db.add(item)
        created += 1

    db.commit()

    return {
        "filename": original_filename,
        "created": created,
        "updated": updated,
        "skipped": skipped,
    }
