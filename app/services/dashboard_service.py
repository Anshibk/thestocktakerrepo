from __future__ import annotations

import uuid
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable, Iterator, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session, aliased

from app.models.category import CategoryGroup, SubCategory
from app.models.entry import Entry
from app.models.item import Item
from app.models.session_inv import InventorySession
from app.models.user import User
from app.models.warehouse import Warehouse


CURRENCY_FORMAT = "₹#,##0.00"

CurrencyFill = PatternFill("solid", fgColor="EFF6FF")
HeaderFill = PatternFill("solid", fgColor="0F172A")
HeaderFont = Font(color="FFFFFF", bold=True)
TitleFont = Font(size=14, bold=True)
SubtitleFont = Font(size=11, color="606C80")
ZebraFill = PatternFill("solid", fgColor="F8FAFC")
ThinBorder = Border(
    left=Side(style="thin", color="CBD5F5"),
    right=Side(style="thin", color="CBD5F5"),
    top=Side(style="thin", color="CBD5F5"),
    bottom=Side(style="thin", color="CBD5F5"),
)


def _decimal_to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):  # noqa: PERF203 - defensive
        return 0.0


def _slugify(value: str) -> str:
    cleaned = [c.lower() if c.isalnum() else "-" for c in value.strip()]
    slug = "".join(cleaned)
    slug = "-".join(filter(None, slug.split("-")))
    return slug or "dashboard-export"


def _safe_sheet_title(value: str) -> str:
    cleaned = "".join(" " if ch in "[]:*?/\\" else ch for ch in value.strip())
    cleaned = cleaned or "Sheet"
    return cleaned[:31]


def _format_decimal_label(value: float | None, decimals: int = 3) -> str:
    if value is None:
        return "0"
    formatted = f"{value:,.{decimals}f}".rstrip("0").rstrip(".")
    return formatted or "0"


def _format_qty_text(value: float | None, unit: str | None) -> str:
    base = _format_decimal_label(_decimal_to_float(value))
    unit_text = (unit or "").strip()
    return f"{base} {unit_text}".strip()


def _format_date_label(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value.date()
    elif isinstance(value, date):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return ""
        try:
            cleaned = text.replace("Z", "+00:00")
            dt = datetime.fromisoformat(cleaned).date()
        except ValueError:
            parts = text.split("-")
            if len(parts) == 3:
                try:
                    dt = date(int(parts[0]), int(parts[1]), int(parts[2][:2]))
                except ValueError:
                    return text
            else:
                return text
    return dt.strftime("%d/%m/%Y")


def _summarise_qty_strings(pairs: list[tuple[float, str | None]]) -> str:
    totals: dict[str, float] = {}
    order: list[str] = []
    for value, unit in pairs:
        unit_key = (unit or "").strip()
        if unit_key not in totals:
            totals[unit_key] = 0.0
            order.append(unit_key)
        totals[unit_key] += _decimal_to_float(value)
    parts = [
        _format_qty_text(totals[key], key if key else None)
        for key in order
        if totals[key]
    ]
    return " + ".join(parts) if parts else "0"


def _stream_entry_rows(query, chunk_size: int = 500) -> Iterator[dict[str, Any]]:
    stream = (
        query.execution_options(stream_results=True)
        .yield_per(chunk_size)
    )
    for row in stream:
        yield {
            "entry_id": row.entry_id,
            "item_id": row.item_id,
            "username": row.username,
            "item_name": row.item_name,
            "category_name": row.category_name,
            "batch": row.batch,
            "mfg": row.mfg,
            "exp": row.exp,
            "qty": _decimal_to_float(row.qty),
            "unit": row.unit,
            "location": row.location,
            "price": None if row.price is None else _decimal_to_float(row.price),
            "line_value": _decimal_to_float(row.line_value),
            "created_at": row.created_at,
        }


def cards(db: Session, user_ids: list[uuid.UUID] | None) -> list[dict[str, Any]]:
    category_rows = (
        db.query(
            CategoryGroup.id.label("group_id"),
            CategoryGroup.name.label("group_name"),
            func.count(func.distinct(SubCategory.id)).label("categories"),
            func.count(func.distinct(Item.id)).label("items"),
        )
        .outerjoin(SubCategory, SubCategory.group_id == CategoryGroup.id)
        .outerjoin(Item, Item.category_id == SubCategory.id)
        .group_by(CategoryGroup.id, CategoryGroup.name)
        .order_by(CategoryGroup.name)
        .all()
    )

    totals_query = (
        db.query(
            CategoryGroup.id.label("group_id"),
            func.count(func.distinct(Entry.item_id)).label("counted"),
            func.sum(
                Entry.qty * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("total_value"),
        )
        .select_from(CategoryGroup)
        .outerjoin(SubCategory, SubCategory.group_id == CategoryGroup.id)
        .outerjoin(Item, Item.category_id == SubCategory.id)
        .outerjoin(Entry, Entry.item_id == Item.id)
    )
    if user_ids:
        totals_query = totals_query.filter(Entry.user_id.in_(user_ids))
    totals = {
        row.group_id: row
        for row in totals_query.group_by(CategoryGroup.id).all()
    }

    results: list[dict[str, Any]] = []
    for row in category_rows:
        metrics = totals.get(row.group_id)
        counted = 0
        total_value = 0.0
        if metrics:
            counted = int(metrics.counted or 0)
            total_value = _decimal_to_float(metrics.total_value)
        results.append(
            {
                "group_name": row.group_name,
                "categories": int(row.categories or 0),
                "items": int(row.items or 0),
                "counted": counted,
                "total_value": total_value,
            }
        )
    return results


def table(db: Session, user_ids: list[uuid.UUID] | None) -> list[dict[str, Any]]:
    query = (
        db.query(
            Entry.item_id.label("item_id"),
            Item.name.label("item_name"),
            Item.unit.label("unit"),
            SubCategory.name.label("category_name"),
            func.count(func.distinct(Entry.batch)).label("batches"),
            func.count(Entry.id).label("entries_logged"),
            func.sum(Entry.qty).label("total_qty"),
            func.sum(
                Entry.qty * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("total_value"),
        )
        .join(Item, Item.id == Entry.item_id)
        .outerjoin(SubCategory, SubCategory.id == Entry.category_id)
    )
    if user_ids:
        query = query.filter(Entry.user_id.in_(user_ids))
    rows = (
        query.group_by(Entry.item_id, Item.name, Item.unit, SubCategory.name)
        .order_by(Item.name)
        .all()
    )

    data: list[dict[str, Any]] = []
    for row in rows:
        data.append(
            {
                "item_id": row.item_id,
                "item_name": row.item_name,
                "unit": row.unit,
                "category_name": row.category_name,
                "batches": int(row.batches or 0),
                "entries_logged": int(row.entries_logged or 0),
                "total_qty": _decimal_to_float(row.total_qty),
                "total_value": _decimal_to_float(row.total_value),
            }
        )
    return data


def detail(
    db: Session,
    user_ids: list[uuid.UUID] | None,
    item_id: uuid.UUID,
    *,
    limit: int | None = 50,
    offset: int = 0,
) -> dict[str, Any]:
    item_row = (
        db.query(
            Item.id.label("item_id"),
            Item.name.label("item_name"),
            SubCategory.name.label("category_name"),
        )
        .outerjoin(SubCategory, Item.category_id == SubCategory.id)
        .filter(Item.id == item_id)
        .one_or_none()
    )
    if not item_row:
        return {"item": None, "entries": []}

    query = (
        db.query(
            Entry.id.label("entry_id"),
            User.username.label("username"),
            Item.name.label("item_name"),
            SubCategory.name.label("category_name"),
            Item.unit.label("unit"),
            Entry.batch,
            Entry.mfg,
            Entry.exp,
            Entry.qty,
            Warehouse.name.label("location"),
            func.coalesce(Entry.price_at_entry, Item.price, 0).label("price"),
            (
                Entry.qty
                * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("line_value"),
            Entry.created_at,
        )
        .join(User, User.id == Entry.user_id)
        .join(Item, Item.id == Entry.item_id)
        .outerjoin(SubCategory, SubCategory.id == Entry.category_id)
        .join(Warehouse, Warehouse.id == Entry.warehouse_id)
        .filter(Entry.item_id == item_id)
    )
    if user_ids:
        query = query.filter(Entry.user_id.in_(user_ids))
    total = query.count()

    ordered_query = query.order_by(Entry.created_at.desc())
    if limit is not None:
        ordered_query = ordered_query.offset(offset).limit(limit)
    rows = ordered_query.all()

    entries: list[dict[str, Any]] = []
    for row in rows:
        qty = _decimal_to_float(row.qty)
        price = _decimal_to_float(row.price)
        line_value = _decimal_to_float(row.line_value)
        entries.append(
            {
                "entry_id": row.entry_id,
                "username": row.username,
                "item_name": row.item_name,
                "category_name": row.category_name,
                "unit": row.unit,
                "batch": row.batch,
                "qty": qty,
                "mfg": row.mfg,
                "exp": row.exp,
                "location": row.location,
                "price": price,
                "line_value": line_value,
                "created_at": row.created_at,
            }
        )

    effective_limit = limit if limit is not None else total
    effective_offset = offset if limit is not None else 0
    has_next = False if limit is None else (effective_offset + len(entries) < total)

    return {
        "item": {
            "item_id": item_row.item_id,
            "item_name": item_row.item_name,
            "category_name": item_row.category_name,
        },
        "entries": entries,
        "total": total,
        "limit": effective_limit,
        "offset": effective_offset,
        "has_next": has_next,
    }


def _style_header_row(sheet, titles: list[str], start_row: int = 1) -> int:
    for idx, title in enumerate(titles, start=1):
        cell = sheet.cell(row=start_row, column=idx)
        cell.value = title
        cell.font = HeaderFont
        cell.fill = HeaderFill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = ThinBorder
    return start_row + 1


def _autofit_columns(sheet, max_column: int, minimum: int = 10) -> None:
    for col_idx in range(1, max_column + 1):
        letter = get_column_letter(col_idx)
        max_length = 0
        for cell in sheet[letter]:
            if cell.value is None:
                continue
            value = cell.value
            if isinstance(value, float):
                text = f"{value:.3f}".rstrip("0").rstrip(".")
            else:
                text = str(value)
            max_length = max(max_length, len(text))
        sheet.column_dimensions[letter].width = max(minimum, min(max_length + 4, 50))


def _normalise_group_name(value: str | None) -> str:
    name = (value or "").strip()
    return name or "Uncategorised"


def _compute_group_and_subcategory_stats(
    summaries: Sequence[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    group_map: dict[str, dict[str, Any]] = {}
    subcategory_map: dict[str, dict[str, Any]] = {}

    for summary in summaries:
        group_name = _normalise_group_name(summary.get("group_name"))
        subcategory_name = (summary.get("category_name") or "").strip() or "Uncategorised"

        group_entry = group_map.setdefault(
            group_name,
            {
                "subcategories": set(),
                "items": 0,
                "total_qty": 0.0,
                "total_value": 0.0,
            },
        )
        group_entry["items"] += 1
        group_entry["total_qty"] += _decimal_to_float(summary.get("total_qty"))
        group_entry["total_value"] += _decimal_to_float(summary.get("total_value"))
        group_entry["subcategories"].add(subcategory_name)

        sub_entry = subcategory_map.setdefault(
            subcategory_name,
            {
                "group_name": group_name,
                "total_qty": 0.0,
                "total_value": 0.0,
            },
        )
        sub_entry["total_qty"] += _decimal_to_float(summary.get("total_qty"))
        sub_entry["total_value"] += _decimal_to_float(summary.get("total_value"))

    group_summary = [
        {
            "group_name": name,
            "subcategories_count": len(data["subcategories"]),
            "items_count": data["items"],
            "total_qty": data["total_qty"],
            "total_value": data["total_value"],
        }
        for name, data in group_map.items()
    ]
    group_summary.sort(key=lambda row: row["group_name"].lower())

    subcategory_summary = [
        {
            "subcategory_name": name,
            "group_name": data["group_name"],
            "total_qty": data["total_qty"],
            "total_value": data["total_value"],
        }
        for name, data in subcategory_map.items()
    ]
    subcategory_summary.sort(
        key=lambda row: (-row["total_value"], row["subcategory_name"].lower())
    )

    return group_summary, subcategory_summary


def _build_master_item_sheet(
    workbook: Workbook, summaries: Sequence[dict[str, Any]], generated_label: str
) -> None:
    sheet = workbook.active
    sheet.title = "Master_Item_wise_report"

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Master Item-wise Inventory Report"
    title_cell.font = TitleFont
    title_cell.alignment = Alignment(horizontal="left")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f"Generated on {generated_label}"
    subtitle_cell.font = SubtitleFont
    subtitle_cell.alignment = Alignment(horizontal="left")

    headers = [
        "#",
        "Item Name",
        "Sub category",
        "Batches",
        "Entries logged",
        "Total Quantity",
        "Total Value",
    ]
    header_row = _style_header_row(sheet, headers, start_row=4)
    sheet.freeze_panes = sheet[f"A{header_row}"]

    if not summaries:
        empty_cell = sheet.cell(row=header_row, column=1)
        empty_cell.value = "No items available"
        empty_cell.alignment = Alignment(horizontal="left")
        empty_cell.font = Font(color="64748B")
        _autofit_columns(sheet, 7)
        return

    for index, summary in enumerate(summaries, start=1):
        row_idx = header_row + index - 1
        zebra = index % 2 == 0

        category_display = summary.get("category_name") or "—"
        cells = [
            (1, index, Alignment(horizontal="center"), None),
            (2, summary.get("item_name"), Alignment(horizontal="left"), None),
            (3, category_display, Alignment(horizontal="left"), None),
            (4, int(summary.get("batches", 0) or 0), Alignment(horizontal="center"), "0"),
            (5, int(summary.get("entries_logged", 0) or 0), Alignment(horizontal="center"), "0"),
            (
                6,
                _format_qty_text(summary.get("total_qty"), summary.get("unit")),
                Alignment(horizontal="right"),
                "@",
            ),
            (7, summary.get("total_value"), Alignment(horizontal="right"), CURRENCY_FORMAT),
        ]

        for column, value, alignment, number_format in cells:
            cell = sheet.cell(row=row_idx, column=column)
            cell.value = value
            cell.alignment = alignment
            cell.border = ThinBorder
            if zebra:
                cell.fill = ZebraFill
            if number_format:
                cell.number_format = number_format

    _autofit_columns(sheet, 7, minimum=14)


def _build_master_entries_sheet(
    workbook: Workbook,
    detail_rows: Iterable[dict[str, Any]],
    generated_label: str,
) -> None:
    sheet = workbook.create_sheet(title="Master_logged_entries_report")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Master Logged Entries Report"
    title_cell.font = TitleFont
    title_cell.alignment = Alignment(horizontal="left")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f"Generated on {generated_label}"
    subtitle_cell.font = SubtitleFont
    subtitle_cell.alignment = Alignment(horizontal="left")

    headers = [
        "#",
        "Entry Date",
        "User",
        "Item Name",
        "Sub category",
        "Batch",
        "Mfg Date",
        "Exp Date",
        "Quantity",
        "Location",
        "Price",
        "Line Value",
    ]
    header_row = _style_header_row(sheet, headers, start_row=4)
    sheet.freeze_panes = sheet[f"A{header_row}"]

    line_number = 0
    for line_number, entry in enumerate(detail_rows, start=1):
        row_idx = header_row + line_number - 1
        zebra = line_number % 2 == 0
        category_display = entry.get("category_name") or "—"
        location_display = entry.get("location") or "—"
        batch_display = entry.get("batch") or "—"
        cells = [
            (1, line_number, Alignment(horizontal="center"), None),
            (2, _format_date_label(entry.get("created_at")), Alignment(horizontal="center"), None),
            (3, entry.get("username"), Alignment(horizontal="left"), None),
            (4, entry.get("item_name"), Alignment(horizontal="left"), None),
            (5, category_display, Alignment(horizontal="left"), None),
            (6, batch_display, Alignment(horizontal="left"), None),
            (7, entry.get("mfg"), Alignment(horizontal="center"), None),
            (8, entry.get("exp"), Alignment(horizontal="center"), None),
            (
                9,
                _format_qty_text(entry.get("qty"), entry.get("unit")),
                Alignment(horizontal="right"),
                "@",
            ),
            (10, location_display, Alignment(horizontal="left"), None),
            (
                11,
                entry.get("price"),
                Alignment(horizontal="right"),
                CURRENCY_FORMAT if entry.get("price") is not None else None,
            ),
            (12, entry.get("line_value"), Alignment(horizontal="right"), CURRENCY_FORMAT),
        ]

        for column, value, alignment, number_format in cells:
            cell = sheet.cell(row=row_idx, column=column)
            cell.value = value
            cell.alignment = alignment
            cell.border = ThinBorder
            if zebra:
                cell.fill = ZebraFill
            if number_format:
                cell.number_format = number_format

    if line_number == 0:
        empty_cell = sheet.cell(row=header_row, column=1)
        empty_cell.value = "No logged entries available"
        empty_cell.alignment = Alignment(horizontal="left")
        empty_cell.font = Font(color="64748B")

    _autofit_columns(sheet, 12, minimum=14)


def _build_valued_charts_sheet(
    workbook: Workbook,
    group_summary: list[dict[str, Any]],
    subcategory_summary: list[dict[str, Any]],
    generated_label: str,
) -> None:
    sheet = workbook.active
    sheet.title = "valued_charts"
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Inventory Valuation Insights"
    title_cell.font = Font(size=18, bold=True, color="0F172A")
    title_cell.alignment = Alignment(horizontal="left")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f"Generated on {generated_label}"
    subtitle_cell.font = SubtitleFont
    subtitle_cell.alignment = Alignment(horizontal="left")

    total_value = sum(_decimal_to_float(row.get("total_value")) for row in group_summary)
    total_qty = sum(_decimal_to_float(row.get("total_qty")) for row in group_summary)

    top_group_name = "—"
    top_group_value = 0.0
    if group_summary:
      top_group = max(group_summary, key=lambda row: _decimal_to_float(row.get("total_value") or 0))
      top_group_name = top_group.get("group_name") or top_group_name
      top_group_value = _decimal_to_float(top_group.get("total_value"))

    top_group_share = (top_group_value / total_value) if total_value else 0
    avg_value_per_group = (total_value / len(group_summary)) if group_summary else 0

    card_fill = PatternFill("solid", fgColor="E0F2FE")
    value_fill = PatternFill("solid", fgColor="F8FAFC")

    group_count = len(group_summary)
    card_specs: list[tuple[str, Any, str | None, str | None]] = [
        ("Total Inventory Value", total_value, CURRENCY_FORMAT, f"Across {group_count} group(s)" if group_count else "No active groups"),
        ("Average Value per Group", avg_value_per_group, CURRENCY_FORMAT, f"Based on {group_count} group(s)" if group_count else "No active groups"),
        ("Top Group by Value", top_group_name, None, 
         "No valuation data" if not total_value else f"₹{_format_decimal_label(top_group_value, 2)} • {top_group_share:.1%} share"),
    ]

    card_row_label = 4
    card_row_value = 5
    card_row_meta = 6

    for idx, (label, value, number_format, meta) in enumerate(card_specs):
        start_col = 1 + (idx * 4)
        end_col = start_col + 3
        sheet.merge_cells(start_row=card_row_label, start_column=start_col, end_row=card_row_label, end_column=end_col)
        label_cell = sheet.cell(row=card_row_label, column=start_col)
        label_cell.value = label
        label_cell.font = Font(size=11, color="0F172A", bold=True)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.fill = card_fill

        sheet.merge_cells(start_row=card_row_value, start_column=start_col, end_row=card_row_value, end_column=end_col)
        value_cell = sheet.cell(row=card_row_value, column=start_col)
        value_cell.value = value
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell.font = Font(size=18, color="0F172A", bold=True)
        value_cell.fill = value_fill
        if number_format:
            value_cell.number_format = number_format

        sheet.merge_cells(start_row=card_row_meta, start_column=start_col, end_row=card_row_meta, end_column=end_col)
        meta_cell = sheet.cell(row=card_row_meta, column=start_col)
        if meta:
            meta_cell.value = meta
        elif isinstance(value, (int, float)) and number_format == CURRENCY_FORMAT:
            meta_cell.value = f"≈ {_format_decimal_label(_decimal_to_float(value), 2)}"
        else:
            meta_cell.value = ""
        meta_cell.font = Font(size=10, color="64748B")
        meta_cell.alignment = Alignment(horizontal="left", vertical="center")
        meta_cell.fill = value_fill

    table_start = card_row_meta + 2
    headers = [
        "Group Category",
        "Total Value",
        "Total Quantity",
        "Value Share %",
    ]
    header_row = _style_header_row(sheet, headers, start_row=table_start)

    sorted_groups = sorted(
        group_summary,
        key=lambda row: _decimal_to_float(row.get("total_value") or 0),
        reverse=True,
    )

    if sorted_groups:
        for index, group in enumerate(sorted_groups, start=1):
            row_idx = header_row + index - 1
            zebra = index % 2 == 0
            value = _decimal_to_float(group.get("total_value"))
            qty = _decimal_to_float(group.get("total_qty"))
            share_fraction = (value / total_value) if total_value else 0
            cells = [
                (1, group.get("group_name"), Alignment(horizontal="left"), None),
                (2, value, Alignment(horizontal="right"), CURRENCY_FORMAT),
                (3, qty, Alignment(horizontal="right"), "#,##0.00"),
                (4, share_fraction, Alignment(horizontal="right"), "0.0%"),
            ]
            for column, cell_value, alignment, number_format in cells:
                cell = sheet.cell(row=row_idx, column=column)
                cell.value = cell_value
                cell.alignment = alignment
                cell.border = ThinBorder
                if zebra:
                    cell.fill = ZebraFill
                if number_format:
                    cell.number_format = number_format
    else:
        empty_cell = sheet.cell(row=header_row, column=1)
        empty_cell.value = "No valuation data available"
        empty_cell.font = Font(color="64748B")
        empty_cell.alignment = Alignment(horizontal="left")

    group_table_end = header_row + max(len(sorted_groups), 1) - 1

    if sorted_groups:
        value_ref = Reference(sheet, min_col=2, min_row=header_row, max_row=group_table_end)
        qty_ref = Reference(sheet, min_col=3, min_row=header_row, max_row=group_table_end)
        category_ref = Reference(sheet, min_col=1, min_row=header_row, max_row=group_table_end)

        pie = PieChart()
        pie.title = "Value Mix by Group"
        pie.height = 9
        pie.width = 18
        pie.innerRadius = 40
        pie.add_data(value_ref, titles_from_data=False)
        pie.set_categories(category_ref)
        data_labels = DataLabelList()
        data_labels.showPercent = True
        data_labels.showLeaderLines = True
        pie.dataLabels = data_labels
        sheet.add_chart(pie, f"H{header_row - 1}")

        combo = BarChart()
        combo.title = "Value vs Quantity by Group"
        combo.y_axis.title = "Total Value"
        combo.y_axis.number_format = CURRENCY_FORMAT
        combo.y_axis.majorGridlines = None
        combo.add_data(value_ref, titles_from_data=False)
        combo.set_categories(category_ref)
        combo.height = 9
        combo.width = 18

        qty_line = LineChart()
        qty_line.add_data(qty_ref, titles_from_data=False)
        qty_line.y_axis.axId = 200
        qty_line.y_axis.title = "Quantity"
        qty_line.y_axis.number_format = "#,##0.00"
        qty_line.y_axis.crossAx = combo.x_axis.axId
        qty_line.y_axis.majorGridlines = None
        if qty_line.series:
            series = qty_line.series[0]
            series.graphicalProperties.line.solidFill = "0EA5E9"
            series.graphicalProperties.line.width = 18000
        combo += qty_line
        sheet.add_chart(combo, f"H{header_row + 16}")

    sub_start = group_table_end + 3
    sub_headers = [
        "Top Sub Categories",
        "Group",
        "Total Value",
        "Quantity",
        "Share %",
    ]
    sub_header_row = _style_header_row(sheet, sub_headers, start_row=sub_start)

    top_subcategories = sorted(
        subcategory_summary,
        key=lambda row: _decimal_to_float(row.get("total_value") or 0),
        reverse=True,
    )[:10]

    if top_subcategories:
        for index, sub in enumerate(top_subcategories, start=1):
            row_idx = sub_header_row + index - 1
            zebra = index % 2 == 0
            sub_value = _decimal_to_float(sub.get("total_value"))
            sub_qty = _decimal_to_float(sub.get("total_qty"))
            share_fraction = (sub_value / total_value) if total_value else 0
            cells = [
                (1, sub.get("subcategory_name"), Alignment(horizontal="left"), None),
                (2, sub.get("group_name"), Alignment(horizontal="left"), None),
                (3, sub_value, Alignment(horizontal="right"), CURRENCY_FORMAT),
                (4, sub_qty, Alignment(horizontal="right"), "#,##0.00"),
                (5, share_fraction, Alignment(horizontal="right"), "0.0%"),
            ]
            for column, cell_value, alignment, number_format in cells:
                cell = sheet.cell(row=row_idx, column=column)
                cell.value = cell_value
                cell.alignment = alignment
                cell.border = ThinBorder
                if zebra:
                    cell.fill = ZebraFill
                if number_format:
                    cell.number_format = number_format
        sub_table_end = sub_header_row + len(top_subcategories) - 1

        bar = BarChart()
        bar.type = "bar"
        bar.title = "Top Sub Categories by Value"
        bar.y_axis.number_format = CURRENCY_FORMAT
        bar.add_data(
            Reference(sheet, min_col=3, min_row=sub_header_row, max_row=sub_table_end),
            titles_from_data=False,
        )
        bar.set_categories(
            Reference(sheet, min_col=1, min_row=sub_header_row, max_row=sub_table_end)
        )
        bar.height = 10
        bar.width = 18
        sheet.add_chart(bar, f"A{sub_table_end + 3}")

        share_range = f"E{sub_header_row}:E{sub_table_end}"
        sheet.conditional_formatting.add(
            share_range,
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="38BDF8"),
        )
    else:
        empty_cell = sheet.cell(row=sub_header_row, column=1)
        empty_cell.value = "No sub category data available"
        empty_cell.font = Font(color="64748B")
        empty_cell.alignment = Alignment(horizontal="left")

    sheet.row_dimensions[card_row_label].height = 26
    sheet.row_dimensions[card_row_value].height = 32
    sheet.row_dimensions[card_row_meta].height = 18

    sheet.freeze_panes = sheet[f"A{header_row}"]
    _autofit_columns(sheet, 12, minimum=14)


def _build_valued_group_summary_sheet(
    workbook: Workbook,
    group_summary: list[dict[str, Any]],
    generated_label: str,
) -> None:
    sheet = workbook.create_sheet(title="valuated_group_summary")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Valued Group Summary"
    title_cell.font = TitleFont
    title_cell.alignment = Alignment(horizontal="left")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f"Generated on {generated_label}"
    subtitle_cell.font = SubtitleFont
    subtitle_cell.alignment = Alignment(horizontal="left")

    headers = [
        "Group Category",
        "Sub categories",
        "Items",
        "Counted",
        "Total Value",
    ]
    header_row = _style_header_row(sheet, headers, start_row=4)
    sheet.freeze_panes = sheet[f"A{header_row}"]

    totals = {"subs": 0, "items": 0, "qty": 0.0, "value": 0.0}

    if group_summary:
        for index, group in enumerate(group_summary, start=1):
            row_idx = header_row + index - 1
            zebra = index % 2 == 0
            subs = int(group.get("subcategories_count", 0) or 0)
            items = int(group.get("items_count", 0) or 0)
            qty = _decimal_to_float(group.get("total_qty"))
            value = _decimal_to_float(group.get("total_value"))
            totals["subs"] += subs
            totals["items"] += items
            totals["qty"] += qty
            totals["value"] += value

            cells = [
                (1, group.get("group_name"), Alignment(horizontal="left"), None),
                (2, subs, Alignment(horizontal="center"), "0"),
                (3, items, Alignment(horizontal="center"), "0"),
                (4, qty, Alignment(horizontal="right"), "#,##0.00"),
                (5, value, Alignment(horizontal="right"), CURRENCY_FORMAT),
            ]

            for column, value_cell, alignment, number_format in cells:
                cell = sheet.cell(row=row_idx, column=column)
                cell.value = value_cell
                cell.alignment = alignment
                cell.border = ThinBorder
                if zebra:
                    cell.fill = ZebraFill
                if number_format:
                    cell.number_format = number_format
    else:
        empty_cell = sheet.cell(row=header_row, column=1)
        empty_cell.value = "No valuation data available"
        empty_cell.font = Font(color="64748B")
        empty_cell.alignment = Alignment(horizontal="left")

    totals_row = header_row + max(len(group_summary), 1)
    label_cell = sheet.cell(row=totals_row, column=1)
    label_cell.value = "Totals"
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="right")
    label_cell.border = ThinBorder

    subs_cell = sheet.cell(row=totals_row, column=2)
    subs_cell.value = totals["subs"]
    subs_cell.font = Font(bold=True)
    subs_cell.alignment = Alignment(horizontal="center")
    subs_cell.border = ThinBorder
    subs_cell.number_format = "0"

    items_cell = sheet.cell(row=totals_row, column=3)
    items_cell.value = totals["items"]
    items_cell.font = Font(bold=True)
    items_cell.alignment = Alignment(horizontal="center")
    items_cell.border = ThinBorder
    items_cell.number_format = "0"

    qty_cell = sheet.cell(row=totals_row, column=4)
    qty_cell.value = totals["qty"]
    qty_cell.font = Font(bold=True)
    qty_cell.alignment = Alignment(horizontal="right")
    qty_cell.border = ThinBorder
    qty_cell.number_format = "#,##0.00"

    value_cell = sheet.cell(row=totals_row, column=5)
    value_cell.value = totals["value"]
    value_cell.font = Font(bold=True)
    value_cell.alignment = Alignment(horizontal="right")
    value_cell.border = ThinBorder
    value_cell.number_format = CURRENCY_FORMAT
    value_cell.fill = CurrencyFill

    _autofit_columns(sheet, 5, minimum=14)


def _build_valued_item_sheet(
    workbook: Workbook,
    summaries: Sequence[dict[str, Any]],
    generated_label: str,
) -> None:
    sheet = workbook.create_sheet(title="Valued_Item_wise_report")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Valued Item-wise Report"
    title_cell.font = TitleFont
    title_cell.alignment = Alignment(horizontal="left")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f"Generated on {generated_label}"
    subtitle_cell.font = SubtitleFont
    subtitle_cell.alignment = Alignment(horizontal="left")

    headers = [
        "#",
        "Item Name",
        "Sub category",
        "Batches",
        "Entries logged",
        "Total Quantity",
        "Total Value",
    ]
    header_row = _style_header_row(sheet, headers, start_row=4)
    sheet.freeze_panes = sheet[f"A{header_row}"]

    if not summaries:
        empty_cell = sheet.cell(row=header_row, column=1)
        empty_cell.value = "No valuation data available"
        empty_cell.font = Font(color="64748B")
        empty_cell.alignment = Alignment(horizontal="left")
        _autofit_columns(sheet, 7)
        return

    for index, summary in enumerate(summaries, start=1):
        row_idx = header_row + index - 1
        zebra = index % 2 == 0

        category_display = summary.get("category_name") or "—"
        cells = [
            (1, index, Alignment(horizontal="center"), None),
            (2, summary.get("item_name"), Alignment(horizontal="left"), None),
            (3, category_display, Alignment(horizontal="left"), None),
            (4, int(summary.get("batches", 0) or 0), Alignment(horizontal="center"), "0"),
            (5, int(summary.get("entries_logged", 0) or 0), Alignment(horizontal="center"), "0"),
            (
                6,
                _format_qty_text(summary.get("total_qty"), summary.get("unit")),
                Alignment(horizontal="right"),
                "@",
            ),
            (7, summary.get("total_value"), Alignment(horizontal="right"), CURRENCY_FORMAT),
        ]

        for column, value, alignment, number_format in cells:
            cell = sheet.cell(row=row_idx, column=column)
            cell.value = value
            cell.alignment = alignment
            cell.border = ThinBorder
            if zebra:
                cell.fill = ZebraFill
            if number_format:
                cell.number_format = number_format

    last_row = header_row + len(summaries) - 1
    if last_row >= header_row:
        qty_range = f"F{header_row}:F{last_row}"
        value_range = f"G{header_row}:G{last_row}"
        sheet.conditional_formatting.add(
            qty_range,
            DataBarRule(start_type="min", end_type="max", color="38BDF8"),
        )
        sheet.conditional_formatting.add(
            value_range,
            DataBarRule(start_type="min", end_type="max", color="0EA5E9"),
        )

    _autofit_columns(sheet, 7, minimum=14)


def _build_valued_entries_sheet(
    workbook: Workbook,
    detail_rows: Iterable[dict[str, Any]],
    generated_label: str,
) -> None:
    sheet = workbook.create_sheet(title="Valued_logged_entries_report")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "Valued Logged Entries Report"
    title_cell.font = TitleFont
    title_cell.alignment = Alignment(horizontal="left")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f"Generated on {generated_label}"
    subtitle_cell.font = SubtitleFont
    subtitle_cell.alignment = Alignment(horizontal="left")

    headers = [
        "#",
        "Entry Date",
        "User",
        "Item Name",
        "Sub category",
        "Batch",
        "Mfg Date",
        "Exp Date",
        "Quantity",
        "Location",
        "Price",
        "Line Value",
    ]
    header_row = _style_header_row(sheet, headers, start_row=4)
    sheet.freeze_panes = sheet[f"A{header_row}"]

    line_number = 0
    for line_number, entry in enumerate(detail_rows, start=1):
        row_idx = header_row + line_number - 1
        zebra = line_number % 2 == 0
        category_display = entry.get("category_name") or "—"
        location_display = entry.get("location") or "—"
        batch_display = entry.get("batch") or "—"
        cells = [
            (1, line_number, Alignment(horizontal="center"), None),
                (2, _format_date_label(entry.get("created_at")), Alignment(horizontal="center"), None),
                (3, entry.get("username"), Alignment(horizontal="left"), None),
                (4, entry.get("item_name"), Alignment(horizontal="left"), None),
                (5, category_display, Alignment(horizontal="left"), None),
                (6, batch_display, Alignment(horizontal="left"), None),
                (7, entry.get("mfg"), Alignment(horizontal="center"), None),
                (8, entry.get("exp"), Alignment(horizontal="center"), None),
                (
                    9,
                    _format_qty_text(entry.get("qty"), entry.get("unit")),
                    Alignment(horizontal="right"),
                    "@",
                ),
                (10, location_display, Alignment(horizontal="left"), None),
                (
                    11,
                    entry.get("price"),
                    Alignment(horizontal="right"),
                    CURRENCY_FORMAT if entry.get("price") is not None else None,
                ),
                (12, entry.get("line_value"), Alignment(horizontal="right"), CURRENCY_FORMAT),
            ]

        for column, value, alignment, number_format in cells:
            cell = sheet.cell(row=row_idx, column=column)
            cell.value = value
            cell.alignment = alignment
            cell.border = ThinBorder
            if zebra:
                cell.fill = ZebraFill
            if number_format:
                cell.number_format = number_format

    if line_number == 0:
        empty_cell = sheet.cell(row=header_row, column=1)
        empty_cell.value = "No valuation entries available"
        empty_cell.font = Font(color="64748B")
        empty_cell.alignment = Alignment(horizontal="left")

    _autofit_columns(sheet, 12, minimum=14)

def export_detail(
    detail_data: dict[str, Any],
    summary_rows: list[dict[str, Any]] | None = None,
) -> tuple[BytesIO, str]:
    workbook = Workbook()

    item = detail_data.get("item") or {}
    entries = detail_data.get("entries") or []

    item_name = item.get("item_name") or "Inventory Item"
    category_name = item.get("category_name") or "Uncategorised"
    item_id = item.get("item_id")

    summary_row: dict[str, Any] | None = None
    if summary_rows:
        if item_id is not None:
            for row in summary_rows:
                if row.get("item_id") == item_id:
                    summary_row = row
                    break
        if summary_row is None:
            for row in summary_rows:
                if row.get("item_name") == item_name:
                    summary_row = row
                    break

    if summary_row:
        total_qty = _decimal_to_float(summary_row.get("total_qty"))
        total_value = _decimal_to_float(summary_row.get("total_value"))
        batches = int(summary_row.get("batches", 0) or 0)
        unit = summary_row.get("unit")
        category_name = summary_row.get("category_name") or category_name
    else:
        batches = len({entry.get("batch") for entry in entries if entry.get("batch")})
        total_qty = sum(_decimal_to_float(entry.get("qty")) for entry in entries)
        total_value = sum(
            _decimal_to_float(entry.get("line_value")) for entry in entries
        )
        unit = None
        for entry in entries:
            if entry.get("unit"):
                unit = entry.get("unit")
                break

    qty_total_pairs: list[tuple[float, str | None]] = [
        (_decimal_to_float(entry.get("qty")), entry.get("unit")) for entry in entries
    ]
    qty_total_text = (
        _summarise_qty_strings(qty_total_pairs)
        if qty_total_pairs
        else _format_qty_text(total_qty, unit)
    )

    sheet = workbook.active
    sheet.title = _safe_sheet_title(f"{item_name}_Summary")

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    header_cell = sheet.cell(row=1, column=1)
    header_cell.value = item_name
    header_cell.font = Font(size=16, bold=True, color="1F2937")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f"Category: {category_name or '—'}"
    subtitle_cell.font = SubtitleFont
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
    meta_cell = sheet.cell(row=3, column=1)
    meta_cell.value = f"Exported On: {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}"
    meta_cell.font = Font(size=10, color="64748B")
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")

    entries_logged = len(entries)
    summary_pairs = [
        ("Total Qty Counted", qty_total_text, None),
        ("Total Inventory Value", total_value, CURRENCY_FORMAT),
        ("Batches", str(batches), None),
        ("Entries logged", str(entries_logged), None),
    ]

    summary_row_index = 5
    for block_index, (label, value, number_format) in enumerate(summary_pairs):
        column = 1 + (block_index * 2)
        label_cell = sheet.cell(row=summary_row_index, column=column)
        label_cell.value = label
        label_cell.font = Font(size=11, bold=True, color="334155")
        label_cell.fill = PatternFill("solid", fgColor="E2E8F0")
        label_cell.alignment = Alignment(horizontal="left")
        label_cell.border = ThinBorder

        value_cell = sheet.cell(row=summary_row_index, column=column + 1)
        value_cell.value = value
        value_cell.font = Font(size=11, color="0F172A")
        value_cell.alignment = Alignment(horizontal="left")
        value_cell.border = ThinBorder
        if number_format:
            value_cell.number_format = number_format
        else:
            value_cell.number_format = "@"

    headers = [
        "#",
        "Entry Date",
        "User",
        "Batch",
        "Mfg",
        "Exp",
        "Quantity",
        "Location",
        "Price",
        "Line Value",
    ]

    header_row = summary_row_index + 2
    body_start = _style_header_row(sheet, headers, start_row=header_row)
    sheet.freeze_panes = sheet["A" + str(body_start)]

    for index, entry in enumerate(entries, start=1):
        row_index = body_start + index - 1
        zebra = index % 2 == 0
        entry_date = _format_date_label(entry.get("created_at"))
        qty_text = _format_qty_text(entry.get("qty"), entry.get("unit"))
        price = entry.get("price")
        line_value = entry.get("line_value")

        cells = [
            (1, index, Alignment(horizontal="center"), None),
            (2, entry_date, Alignment(horizontal="center"), None),
            (3, entry.get("username"), Alignment(horizontal="left"), None),
            (4, entry.get("batch"), Alignment(horizontal="left"), None),
            (5, entry.get("mfg"), Alignment(horizontal="center"), None),
            (6, entry.get("exp"), Alignment(horizontal="center"), None),
            (7, qty_text, Alignment(horizontal="right"), "@"),
            (8, entry.get("location"), Alignment(horizontal="left"), None),
            (
                9,
                price,
                Alignment(horizontal="right"),
                CURRENCY_FORMAT if price is not None else None,
            ),
            (10, line_value, Alignment(horizontal="right"), CURRENCY_FORMAT),
        ]

        for column, value, alignment, number_format in cells:
            cell = sheet.cell(row=row_index, column=column)
            cell.value = value
            cell.alignment = alignment
            cell.border = ThinBorder
            if zebra:
                cell.fill = ZebraFill
            if number_format:
                cell.number_format = number_format

    _autofit_columns(sheet, 10)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"{_slugify(str(item_name))}_summary.xlsx"
    return stream, filename
def export_dashboard(
    db: Session,
    user_ids: list[uuid.UUID] | None,
    *,
    include_master_items: bool,
) -> tuple[BytesIO, str]:
    session_rows = (
        db.query(InventorySession.id, InventorySession.code)
        .filter(InventorySession.status == "active")
        .order_by(InventorySession.name)
        .all()
    )
    session_ids = [row.id for row in session_rows]

    item_category = aliased(SubCategory)
    item_group = aliased(CategoryGroup)
    base_items = (
        db.query(
            Item.id.label("item_id"),
            Item.name.label("item_name"),
            Item.unit.label("unit"),
            item_category.name.label("category_name"),
            item_group.name.label("group_name"),
        )
        .outerjoin(item_category, item_category.id == Item.category_id)
        .outerjoin(item_group, item_group.id == item_category.group_id)
        .order_by(Item.name)
        .all()
    )

    summary_category = aliased(SubCategory)
    summary_group = aliased(CategoryGroup)
    aggregated_query = (
        db.query(
            Entry.item_id.label("item_id"),
            Item.name.label("item_name"),
            Item.unit.label("unit"),
            summary_category.name.label("category_name"),
            summary_group.name.label("group_name"),
            func.count(func.distinct(Entry.batch)).label("batches"),
            func.count(Entry.id).label("entries_logged"),
            func.sum(Entry.qty).label("total_qty"),
            func.sum(
                Entry.qty * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("total_value"),
        )
        .join(Item, Item.id == Entry.item_id)
        .outerjoin(summary_category, summary_category.id == Item.category_id)
        .outerjoin(summary_group, summary_group.id == summary_category.group_id)
    )
    if user_ids:
        aggregated_query = aggregated_query.filter(Entry.user_id.in_(user_ids))
    if session_ids:
        aggregated_query = aggregated_query.filter(Entry.session_id.in_(session_ids))
    aggregated_rows = (
        aggregated_query.group_by(
            Entry.item_id,
            Item.name,
            Item.unit,
            summary_category.name,
            summary_group.name,
        )
        .order_by(Item.name)
        .all()
    )

    aggregated_map: dict[uuid.UUID, dict[str, Any]] = {}
    for row in aggregated_rows:
        aggregated_map[row.item_id] = {
            "item_id": row.item_id,
            "item_name": row.item_name,
            "category_name": row.category_name,
            "group_name": row.group_name,
            "unit": row.unit,
            "batches": int(row.batches or 0),
            "entries_logged": int(row.entries_logged or 0),
            "total_qty": _decimal_to_float(row.total_qty),
            "total_value": _decimal_to_float(row.total_value),
            "has_entries": True,
        }

    entry_category = aliased(SubCategory)
    item_category_detail = aliased(SubCategory)
    detail_query = (
        db.query(
            Entry.id.label("entry_id"),
            Entry.item_id.label("item_id"),
            User.username.label("username"),
            Item.name.label("item_name"),
            func.coalesce(entry_category.name, item_category_detail.name).label(
                "category_name"
            ),
            Entry.batch,
            Entry.mfg,
            Entry.exp,
            Entry.qty,
            Warehouse.name.label("location"),
            func.coalesce(Entry.price_at_entry, Item.price).label("price"),
            (
                Entry.qty
                * func.coalesce(Entry.price_at_entry, Item.price, 0)
            ).label("line_value"),
            Item.unit.label("unit"),
            Entry.created_at.label("created_at"),
        )
        .join(User, User.id == Entry.user_id)
        .join(Item, Item.id == Entry.item_id)
        .outerjoin(entry_category, entry_category.id == Entry.category_id)
        .outerjoin(item_category_detail, item_category_detail.id == Item.category_id)
        .join(Warehouse, Warehouse.id == Entry.warehouse_id)
    )
    if user_ids:
        detail_query = detail_query.filter(Entry.user_id.in_(user_ids))
    if session_ids:
        detail_query = detail_query.filter(Entry.session_id.in_(session_ids))
    detail_query = detail_query.order_by(Item.name, Entry.created_at)

    items_data: dict[uuid.UUID, dict[str, Any]] = {}
    for row in base_items:
        summary = {
            "item_id": row.item_id,
            "item_name": row.item_name,
            "category_name": row.category_name,
            "group_name": row.group_name,
            "unit": row.unit,
            "batches": 0,
            "entries_logged": 0,
            "total_qty": 0.0,
            "total_value": 0.0,
            "has_entries": False,
        }
        aggregated = aggregated_map.get(row.item_id)
        if aggregated:
            summary.update(
                {
                    "batches": aggregated["batches"],
                    "entries_logged": aggregated["entries_logged"],
                    "total_qty": aggregated["total_qty"],
                    "total_value": aggregated["total_value"],
                    "has_entries": True,
                }
            )
            if not summary["category_name"] and aggregated["category_name"]:
                summary["category_name"] = aggregated["category_name"]
            if not summary["group_name"] and aggregated.get("group_name"):
                summary["group_name"] = aggregated.get("group_name")
        items_data[row.item_id] = summary

    for aggregated in aggregated_map.values():
        if aggregated["item_id"] not in items_data:
            items_data[aggregated["item_id"]] = aggregated.copy()

    items_sorted = sorted(
        items_data.values(),
        key=lambda item: (
            0 if item.get("has_entries") else 1,
            (item.get("item_name") or "").lower(),
            (item.get("category_name") or "").lower(),
        ),
    )

    for summary in items_sorted:
        summary["entries_logged"] = int(summary.get("entries_logged") or 0)

    generated_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
    download_date = datetime.utcnow().strftime("%d-%b-%Y")

    workbook = Workbook()

    if include_master_items:
        _build_master_item_sheet(workbook, items_sorted, generated_label)
        detail_stream = _stream_entry_rows(detail_query)
        _build_master_entries_sheet(workbook, detail_stream, generated_label)
        filename = f"Export with master items {download_date}.xlsx"
    else:
        valued_items = [summary for summary in items_sorted if summary.get("has_entries")]
        group_summary, subcategory_summary = _compute_group_and_subcategory_stats(
            valued_items
        )
        _build_valued_charts_sheet(
            workbook,
            group_summary,
            subcategory_summary,
            generated_label,
        )
        _build_valued_group_summary_sheet(
            workbook, group_summary, generated_label
        )
        _build_valued_item_sheet(workbook, valued_items, generated_label)
        detail_stream = _stream_entry_rows(detail_query)
        _build_valued_entries_sheet(workbook, detail_stream, generated_label)
        filename = f"Export valuated items {download_date}.xlsx"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return stream, filename
